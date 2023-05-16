#!/usr/bin/env python3.11
# -*- coding: utf-8 -*-

# script header
# Author:           [Hugo Berra Salazar, ]
# Creation date:    [04/07/2023]
# Description:      [Brief description of the purpose of the script]

# import necessary modules
import os

from datetime           import datetime
from openpyxl           import load_workbook, Workbook
from openpyxl.styles    import Font, PatternFill

def xlsx_general_report(rfc):
    # Gets the current date and time
    fecha_actual  = datetime.now().strftime('%m%d%Y-%H%M%S')

    print('... xlsx in process, please wait.')

    # Set the directory paths
    directories = [
        f'{rfc}/Emisor/Ayudas/',
        f'{rfc}/Emisor/Ingresos/',
        f'{rfc}/Emisor/Pagos/',
        f'{rfc}/Emisor/Nomina/',
        f'{rfc}/Receptor/Descuento_Bonificaciones_Devoluciones/',
        f'{rfc}/Receptor/Gastos',
        f'{rfc}/Receptor/Pagos'
    ]

    # Initialize empty list to store file paths
    file_paths = []

    # Get the most recent file from each directory
    for directory in directories:
        if os.path.exists(directory):
            # Get the most recent file in the directory
            files = os.listdir(directory)
            if files:
                try:
                    file_path = max([os.path.join(directory, f) for f in files if f.endswith('.xlsx')], key=os.path.getctime)
                    file_paths.append(file_path)
                except Exception as e:
                    print(f'E1: {directory}: {e}')
                    pass
            else:
                print(f"E2: Can't find the route: {directory}")
        else:
            print(f"E3: Doesn't exist: {directory}")
     
    # Create a new workbook
    nwb = Workbook()

    # Loop through each Excel file and write its contents to a separate sheet
    for file_path in file_paths:
        try:
            # Load workbook, active worksheet, get the sheet name, create a new sheet
            wb = load_workbook(file_path)
            ws = wb.active
            sn = ws.title
            ns = nwb.create_sheet(sn)

            # Copy the data from the old worksheet to the new worksheet
            for row in ws.iter_rows(values_only=True):
                ns.append(row)

            # Change the font color and fill color of the headers
            for cell in ns[1]:
                cell.font = Font(color='FFFFFF')
                cell.fill = PatternFill(
                    start_color='730707', end_color='730707', fill_type='solid')

        except Exception as e:
            print(f'E4: Error while reading book: {e}')
            continue

    # Delete any extra sheets that were created in the new workbook
    while len(nwb.sheetnames) > len(file_paths):
        nwb.remove(nwb.active)

    # Save the new workbook
    try:
        nwb.save(f'{rfc}/reporte-general-{fecha_actual}.xlsx')
    except Exception as e:
        print(f'E5: Error saving the new workbook: {e}')